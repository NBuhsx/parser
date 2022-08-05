import requests
import openpyxl
from Yarche.configure.config import Config
from Yarche.settings.contol import ProxyPull




from imports import logger
from Yarche.build.base import YarcheBase
from Yarche.build.render_cvs import first_sheet_title_category, required




class Soft:
    def __init__(self, config:Config, proxy:ProxyPull):
        self.__proxy_pull = proxy
        self.__adress_config = iter(config.parser.parse_adress)

        self.proxy = next(self.__proxy_pull)
        self.adress = next(self.__adress_config)
        self.config = config

        self.__index_category = 0
        self.__index_product = 0

    
        self._soft = {
            0: {
                "func": self.build_session, 
                "message":{
                    "before": "Получение сессия", 
                    "after": "Сессия получена"}},
            1: {
                "func": self.get_tokens, 
                "message":{
                    "before": "Получение токенов для взаимодействия с API", 
                    "after": "Токены получены"}},
            2:{
                "func": self.get_coordinats, 
                "message":{
                    "before": "Получение координат", 
                    "after": "Координаты полученны"}},
            3:{
                "func": self.set_adress, 
                "message":{
                    "before": "Установка адреса", 
                    "after": "Адрес Установлен"}},
            4:{
                "func": self.get_categories, 
                "message":{
                    "before": "Получение списка категорий", 
                    "after": "Список категорий получен"}},
            5:{
                "func":self.get_product_category,
                "message":{
                    "before":"Получение продуктов по каждой категории",
                    "after":"Все продукты полученны"}}}



    def center(self, itap:int=0):
        while itap <= len(self._soft.keys()):
            if run := self._soft.get(itap):
                logger.info(run['message'].get("before"))
                run['func']()
                logger.info(run['message'].get("after"))
            itap +=1




    def build_session(self):
        self.session = requests.Session()
        self.session.headers = self.config.parser.headers
        self.session.proxies = self.proxy
        self.session.verify = False
        requests.urllib3.disable_warnings()


    def get_tokens(self):
        if response := YarcheBase.response_base_url(session=self.session):
            self.tokens = {}
            for key, value in YarcheBase.get_tokens(response=response):
                if not key or not value:
                    raise ValueError(itap=1)
                self.tokens[key] = value


    def get_coordinats(self):
        if response := YarcheBase.get_coordinats(
                        geo_token=self.tokens['geotoken'], 
                        adress=self.adress,
                        session=self.session):
            self.coorinats = response


    def set_adress(self):
        if response:= YarcheBase.set_adress(
                    verif_token=self.tokens['token'], 
                    adress=self.adress, 
                    coordinats=self.coorinats,
                    session=self.session):
            match response:
                case {'errors':error} if error:
                    message = error[0].get("message")
                case {'data':data} if data["setCurrentAddress"].get("status") == 'confirmed':
                    return True
                case _:
                    return False


    def get_categories(self):
        if response := YarcheBase.get_categories(
                    verif_token=self.tokens['token'],
                    session=self.session):
            match response:
                case {'data': data} if data:
                    self.categories = data.get('categories')
                case _:
                    return False



    def get_product_category(self):
        for index, category in enumerate(
                    self.categories[self.__index_category:], 
                    start=self.__index_category):

            self.__index_category = index

            if category.get('amount') > 0:
                category.update({"products": category.get('products', [])})

                for page, amound in enumerate(
                            (
                                *(99 for _ in range(category.get('amount') // 99)),
                                category.get('amount') % 99, )[self.__index_product:],
                            start=self.__index_product + 1):

                    self.__index_product = page - 1

                    if response := YarcheBase.get_product(
                                        verif_token=self.tokens['token'],
                                        category_id=category.get('id'),
                                        limit=amound,
                                        page=page,
                                        session=self.session):
                        match response:
                            case {'data': data} if data.get('products'):
                                category.get('products').extend(
                                        data['products'].get('list'))



    def result_csv(self):

        wb = openpyxl.Workbook()
        zero_sheet = wb.create_sheet(title='Category', index=0)
        zero_sheet.append(first_sheet_title_category)
        index = 1
        for category in self.categories:
            zero_sheet.append(
                (category.get('id'),
                category.get('name'),))

            if category.get('products'):
                new_sheet = wb.create_sheet(
                    title=category.get('name') if len(category.get('name')) < 31 else " ".join(
                        category.get('name').split(" ")[:-1]),
                    index=index)
                new_sheet.append(
                    required(inside=category, method=dict.keys)
                )
                new_sheet.append(
                    required(inside=category, method=dict.values)
                )
                new_sheet.append(
                    required(inside=category['products'], method=dict.keys)
                )

                for product in category['products']:
                    new_sheet.append(
                        required(inside=product, method=dict.values)
                    )
                index += 1

        wb.save(
            filename=self.config.setting_system.path_dir_result.joinpath(
                f'result_{self.adress}.xlsx'))
     